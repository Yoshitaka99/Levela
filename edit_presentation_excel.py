from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

folder = Path(r"C:\Users\celic\dormswap\outputs\image_excel")
src = folder / "オロチーム5月復活劇.xlsx"
out = folder / "orochi_may_revival_presentation.xlsx"

wb = openpyxl.load_workbook(src)
for name in ["Presentation_Team", "Presentation_Current", "Presentation_Targets", "Presentation_Action"]:
    if name in wb.sheetnames:
        del wb[name]

members = [
    {"name":"芝隼人","type":"通常","apo":45,"seated":19,"seat_rate":0.422,"contract":4,"close_rate":0.211,"apo_close":0.089,"remain":13,"existing_exp":4.6,"add":32,"add_exp":11.2,"total_exp":15.8,"risk":"中"},
    {"name":"田仲由敬","type":"通常","apo":59,"seated":20,"seat_rate":0.339,"contract":5,"close_rate":0.250,"apo_close":0.085,"remain":29,"existing_exp":10.1,"add":36,"add_exp":12.6,"total_exp":22.7,"risk":"低"},
    {"name":"持木玲那","type":"ベンチ","apo":0,"seated":0,"seat_rate":0.0,"contract":0,"close_rate":0.0,"apo_close":0.0,"remain":0,"existing_exp":0,"add":20,"add_exp":7.0,"total_exp":7.0,"risk":"高"},
    {"name":"木原桃香","type":"通常","apo":0,"seated":0,"seat_rate":0.0,"contract":0,"close_rate":0.0,"apo_close":0.0,"remain":0,"existing_exp":0,"add":0,"add_exp":0,"total_exp":0,"risk":"高"},
    {"name":"加藤陸","type":"ベンチ","apo":26,"seated":9,"seat_rate":0.346,"contract":2,"close_rate":0.222,"apo_close":0.077,"remain":14,"existing_exp":4.9,"add":24,"add_exp":8.4,"total_exp":13.3,"risk":"高"},
    {"name":"河上まちこ","type":"ベンチ","apo":5,"seated":2,"seat_rate":0.400,"contract":0,"close_rate":0.0,"apo_close":0.0,"remain":3,"existing_exp":1.0,"add":20,"add_exp":7.0,"total_exp":8.0,"risk":"高"},
    {"name":"五十嵐凌大","type":"ベンチ","apo":22,"seated":15,"seat_rate":0.682,"contract":1,"close_rate":0.067,"apo_close":0.045,"remain":2,"existing_exp":0.7,"add":24,"add_exp":8.4,"total_exp":9.1,"risk":"高"},
    {"name":"早川大貴","type":"通常","apo":14,"seated":4,"seat_rate":0.286,"contract":1,"close_rate":0.250,"apo_close":0.071,"remain":7,"existing_exp":2.4,"add":22,"add_exp":7.7,"total_exp":10.1,"risk":"中"},
    {"name":"笠松佑衣","type":"通常","apo":30,"seated":10,"seat_rate":0.333,"contract":3,"close_rate":0.300,"apo_close":0.100,"remain":18,"existing_exp":6.3,"add":31,"add_exp":10.9,"total_exp":17.2,"risk":"低"},
    {"name":"和佐田舞緒","type":"通常","apo":23,"seated":11,"seat_rate":0.478,"contract":1,"close_rate":0.091,"apo_close":0.043,"remain":10,"existing_exp":3.5,"add":19,"add_exp":6.7,"total_exp":10.2,"risk":"中"},
    {"name":"関口愛里","type":"通常","apo":64,"seated":20,"seat_rate":0.313,"contract":5,"close_rate":0.250,"apo_close":0.078,"remain":35,"existing_exp":12.3,"add":36,"add_exp":12.6,"total_exp":24.9,"risk":"低"},
]

def title(ws, text, cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(1, 1, text)
    c.fill = PatternFill("solid", fgColor="0B3F66")
    c.font = Font(name="Meiryo UI", size=16, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

def header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = PatternFill("solid", fgColor="C8D9F3")
        cell.font = Font(name="Meiryo UI", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style(ws):
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color="D9E2EC")
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Meiryo UI", size=10, bold=cell.font.bold, color=cell.font.color)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for c in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

ws = wb.create_sheet("Presentation_Team", 0)
title(ws, "発表用：チーム数値サマリー", 6)
ws.append(["項目", "数値", "補足", "", "発表メッセージ", "内容"])
header(ws, 2, 6)
rows = [
    ["目標値", 147, "期間全データのチーム目標", "", "主張", "既存アポだけでは届かない。5/25以降の追加アポ設計が勝負。"],
    ["現状合計", 30, "成約17件＋成約予定13件", "", "配分方針", "成約率・着座率・成約数が良い通常メンバーに厚く、ベンチも毎日稼働へ乗せる。"],
    ["進捗", 0.204, "目標147件に対する進捗", "", "アクション", "90分着金を目指し、商談の無駄を削り、流れを定型化してアポ品質を揃える。"],
    ["不足成約", 124, "今回埋めるべき不足数", "", "", ""],
    ["想定成約率", 0.35, "必要アポ逆算に使用", "", "", ""],
    ["必要アポ数", "=ROUNDUP(B6/B7,0)", "124 ÷ 35%", "", "", ""],
    ["現在残アポ", 96, "前提値。元データ集計は91件", "", "", ""],
    ["追加必要アポ", 264, "実務管理値", "", "", ""],
    ["通常メンバー割当", 176, "負担調整後", "", "", ""],
    ["ベンチメンバー割当", 88, "5/25以降毎日最低1件、終盤2件", "", "", ""],
    ["既存アポ期待成約", "=ROUND(B9*B7,1)", "96 × 35%", "", "", ""],
    ["追加アポ期待成約", "=ROUND(B10*B7,1)", "264 × 35%", "", "", ""],
    ["合計期待成約", "=ROUND(B13+B14,1)", "既存＋追加", "", "", ""],
    ["不足見込み", "=ROUND(B6-B15,1)", "35%前提の微差は商談改善で埋める", "", "", ""],
]
for row in rows: ws.append(row)
for cell in ["B4","B7"]: ws[cell].number_format = "0.0%"
ws.column_dimensions["C"].width = 38
ws.column_dimensions["F"].width = 70
style(ws)

ws = wb.create_sheet("Presentation_Current", 1)
title(ws, "発表用：各メンバー 現在の成績", 12)
headers = ["メンバー名","区分","アポ数","着座数","着座率","成約数","成約率(着座)","アポ対成約率","残アポ","既存期待成約","リスク","発表コメント"]
ws.append(headers)
header(ws, 2, 12)
for m in members:
    comment = "強メンバー。追加配分候補" if m["risk"] == "低" else ("ベンチ。毎日稼働へ引き上げ" if m["type"] == "ベンチ" else "商談改善前提で配分")
    ws.append([m["name"],m["type"],m["apo"],m["seated"],m["seat_rate"],m["contract"],m["close_rate"],m["apo_close"],m["remain"],m["existing_exp"],m["risk"],comment])
for r in range(3, 14):
    for c in [5, 7, 8]:
        ws.cell(r, c).number_format = "0.0%"
ws.column_dimensions["L"].width = 36
style(ws)

ws = wb.create_sheet("Presentation_Targets", 2)
title(ws, "発表用：目標達成に必要なメンバー別数値", 14)
headers = ["メンバー名","区分","現在成約","残アポ","既存期待成約","追加アポ必要数","日平均追加アポ","追加期待成約","合計期待成約","必要着座数目安","必要商談改善ポイント","優先度","リスク","コメント"]
ws.append(headers)
header(ws, 2, 14)
for m in members:
    seat_basis = m["seat_rate"] if m["seat_rate"] > 0 else 0.703
    priority = "最優先" if m["add"] >= 31 or m["type"] == "ベンチ" else ("高" if m["add"] >= 20 else "通常")
    improve = "90分着金の型へ統一" if m["type"] == "通常" else "ロープレ・最終確認後に段階投入"
    ws.append([m["name"],m["type"],m["contract"],m["remain"],m["existing_exp"],m["add"],round(m["add"]/18,1),m["add_exp"],m["total_exp"],round(m["add"]*seat_basis,1),improve,priority,m["risk"],"通常負担を下げるためベンチも最低1件/日" if m["type"] == "ベンチ" else "均しペースで追加アポを消化"])
ws.column_dimensions["K"].width = 28
ws.column_dimensions["N"].width = 38
style(ws)

ws = wb.create_sheet("Presentation_Action", 3)
title(ws, "発表用：アクションプラン", 7)
headers = ["テーマ","目的","具体アクション","提出物","期限","担当","完了基準"]
ws.append(headers)
header(ws, 2, 7)
actions = [
    ["90分着金を目指す","連アポを可能にする","自分の商談とオロさんの商談を比較し、無駄な説明・確認・クロージングのズレを整理","比較メモ＋削る項目リスト","即日","田仲さん","90分で着金まで持っていく型が言語化されている"],
    ["商談の流れを定型化","成約率35%の再現性を上げる","導入→課題確認→提案→決済導線→着金までの標準トークを作る","標準フロー1枚","5/25まで","田仲さん＋オロさん","全員が同じ流れでアポに入れる"],
    ["アポ品質を揃える","連アポでも成約率を落とさない","各アポ前にヒアリング項目・決済導線・想定反論を揃える","アポ前チェックリスト","毎日","各メンバー","商談前準備が5分以内で完了"],
    ["強メンバーへ追加配分","短期で数字を作る","田仲・関口・笠松・芝に厚め、早川・和佐田は改善前提で配分","追加アポ消化表","毎日更新","田仲さん","日別アポ計画との差分ゼロ"],
    ["ベンチ復活","通常メンバーの負担を下げる","5/25以降、ベンチも毎日最低1件。終盤は2件に増やす","ベンチ別ロープレ記録","5/25開始","オロさん確認","許可済みメンバーから順次アポ投入"],
    ["数値レビュー","計画倒れを防ぐ","アポ数・着座・成約・着金を毎日確認し、不足分を翌日に再配分","日次差分表","毎日終業前","田仲さん","不足分が翌日の配分に反映済み"],
]
for a in actions:
    ws.append(a)
for col in ["B","C","D","G"]:
    ws.column_dimensions[col].width = 38
for r in range(3, 9):
    ws.row_dimensions[r].height = 56
style(ws)

wb.save(out)
print(out)
