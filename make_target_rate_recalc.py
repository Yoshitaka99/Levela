from pathlib import Path
import zipfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

BASE = Path(r"C:\Users\celic\dormswap")
XLSX = BASE / r"outputs\image_excel\orochi_may_revival_presentation.xlsx"
OUT = BASE / r"outputs\revival_deck_images_v2"
OUT.mkdir(parents=True, exist_ok=True)

members = [
    {"name":"芝隼人","type":"通常","remain":13,"add":29,"rate":0.50},
    {"name":"田仲由敬","type":"通常","remain":29,"add":32,"rate":0.50},
    {"name":"持木玲那","type":"ベンチ","remain":0,"add":20,"rate":0.40},
    {"name":"木原桃香","type":"通常","remain":0,"add":0,"rate":0.40},
    {"name":"加藤陸","type":"ベンチ","remain":14,"add":24,"rate":0.40},
    {"name":"河上まちこ","type":"ベンチ","remain":3,"add":20,"rate":0.40},
    {"name":"五十嵐凌大","type":"ベンチ","remain":2,"add":24,"rate":0.40},
    {"name":"早川大貴","type":"通常","remain":7,"add":20,"rate":0.45},
    {"name":"笠松佑衣","type":"通常","remain":18,"add":28,"rate":0.40},
    {"name":"和佐田舞緒","type":"通常","remain":10,"add":17,"rate":0.40},
    {"name":"関口愛里","type":"通常","remain":35,"add":32,"rate":0.45},
]
for m in members:
    m["seat"] = round(m["add"] * 0.75, 1)
    m["exist_exp"] = round(m["remain"] * 0.75 * m["rate"], 1)
    m["add_exp"] = round(m["add"] * 0.75 * m["rate"], 1)
    m["total_exp"] = round(m["exist_exp"] + m["add_exp"], 1)

dates = ["5/25","5/26","5/27","5/28","5/29","5/30","5/31","6/1","6/2","6/3","6/4","6/5","6/6","6/7","6/8","6/9","6/10","6/11"]
schedule = {
    "芝隼人":[2]*11+[1]*7,
    "田仲由敬":[2]*14+[1]*4,
    "持木玲那":[1]*16+[2,2],
    "木原桃香":[0]*18,
    "加藤陸":[1]*12+[2]*6,
    "河上まちこ":[1]*16+[2,2],
    "五十嵐凌大":[1]*12+[2]*6,
    "早川大貴":[2,2]+[1]*16,
    "笠松佑衣":[2]*10+[1]*8,
    "和佐田舞緒":[1]*17+[0],
    "関口愛里":[2]*14+[1]*4,
}
daily = [sum(schedule[m["name"]][i] for m in members) for i in range(18)]

# Excel update
wb = openpyxl.load_workbook(XLSX)
for s in ["Presentation_TargetRates", "Presentation_Schedule_TargetRates"]:
    if s in wb.sheetnames:
        del wb[s]

def title(ws, text, cols):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=cols)
    c=ws.cell(1,1,text)
    c.fill=PatternFill("solid",fgColor="0B3F66")
    c.font=Font(name="Meiryo UI",size=16,bold=True,color="FFFFFF")
    c.alignment=Alignment(horizontal="center",vertical="center")
def header(ws,row,cols):
    for c in range(1,cols+1):
        cell=ws.cell(row,c)
        cell.fill=PatternFill("solid",fgColor="C8D9F3")
        cell.font=Font(name="Meiryo UI",bold=True)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def style(ws):
    ws.sheet_view.showGridLines=False
    thin=Side(style="thin",color="D9E2EC")
    for row in ws.iter_rows():
        for cell in row:
            cell.font=Font(name="Meiryo UI",size=10,bold=cell.font.bold,color=cell.font.color)
            cell.alignment=Alignment(vertical="center",wrap_text=True)
            cell.border=Border(top=thin,bottom=thin,left=thin,right=thin)
    for c in range(1,ws.max_column+1):
        ws.column_dimensions[get_column_letter(c)].width=17

ws=wb.create_sheet("Presentation_TargetRates", 6)
title(ws,"発表用：目標成約率別 再計算",12)
ws.append(["メンバー","区分","目標成約率","残アポ","既存期待成約","追加アポ","必要着座(75%)","追加期待成約","合計期待成約","日平均","優先度","コメント"])
header(ws,2,12)
for m in members:
    priority = "最優先" if m["add"] >= 28 or m["type"] == "ベンチ" else ("高" if m["add"] >= 20 else "通常")
    comment = "芝・田仲50%" if m["rate"] == 0.5 else ("関口・早川45%" if m["rate"] == 0.45 else "その他40%")
    ws.append([m["name"],m["type"],m["rate"],m["remain"],m["exist_exp"],m["add"],m["seat"],m["add_exp"],m["total_exp"],round(m["add"]/18,1),priority,comment])
for r in range(3,14):
    ws.cell(r,3).number_format="0%"
style(ws)

ws=wb.create_sheet("Presentation_Schedule_TargetRates", 7)
title(ws,"発表用：目標成約率別 日別追加アポスケジュール",21)
ws.append(["メンバー","区分"]+dates+["合計"])
header(ws,2,21)
for m in members:
    vals=schedule[m["name"]]
    ws.append([m["name"],m["type"]]+vals+[sum(vals)])
ws.append(["日計",""]+daily+[sum(daily)])
style(ws)
wb.save(XLSX)

# Image helpers
W,H=1920,1080
NAVY="#0B3F66"; BG="#F7F9FC"; PANEL="#FFFFFF"; INK="#172033"; MUTED="#607086"; LINE="#D9E2EC"
BLUE="#2D7FF9"; ORANGE="#F59E0B"; RED="#E23B3B"; GREEN="#31A66A"; PALE="#EEF5FF"; WHITE="#FFFFFF"; YELLOW="#FFF2CC"
FONT=r"C:\Windows\Fonts\meiryo.ttc"; FONT_B=r"C:\Windows\Fonts\meiryob.ttc"
def f(s,b=False): return ImageFont.truetype(FONT_B if b else FONT,s)
def text(d,xy,s,size=24,fill=INK,bold=False,anchor=None,align="left",spacing=5):
    d.multiline_text(xy,s,font=f(size,bold),fill=fill,anchor=anchor,align=align,spacing=spacing)
def canvas(title,sub=""):
    im=Image.new("RGB",(W,H),BG); d=ImageDraw.Draw(im)
    d.rectangle([0,0,W,96],fill=NAVY)
    text(d,(64,48),title,34,WHITE,True,"lm")
    if sub: text(d,(W-64,50),sub,20,"#DDEBFF",False,"rm")
    return im,d
def card(d,box,title=None,fill=PANEL):
    d.rounded_rectangle(box,radius=18,fill=fill,outline=LINE,width=2)
    if title: text(d,(box[0]+28,box[1]+28),title,24,MUTED,True)
def metric(d,box,label,value,note="",color=BLUE):
    card(d,box); x1,y1,x2,y2=box
    text(d,(x1+24,y1+24),label,22,MUTED,True)
    text(d,(x1+24,y1+75),value,50,color,True)
    if note: text(d,(x1+24,y2-36),note,19,MUTED)
def bar(d,x,y,w,h,val,maxv,color,right=""):
    d.rounded_rectangle([x,y,x+w,y+h],radius=h//2,fill="#E7EEF6")
    bw=int(w*val/maxv) if maxv else 0
    if bw: d.rounded_rectangle([x,y,x+bw,y+h],radius=h//2,fill=color)
    if right: text(d,(x+w+14,y+h/2),right,18,INK,True,"lm")
def save(im,i,name):
    p=OUT/f"{i:02d}_{name}.png"; im.save(p,quality=95); return p

total_add=sum(m["add"] for m in members)
exist=sum(m["exist_exp"] for m in members)
add_exp=sum(m["add_exp"] for m in members)
total_exp=sum(m["total_exp"] for m in members)

# 14 summary
im,d=canvas("目標成約率別 再計算サマリー","芝・田仲50% / 関口・早川45% / その他40%")
metric(d,(90,170,430,365),"追加アポ","246件","264件から通常負担を圧縮",BLUE)
metric(d,(470,170,810,365),"既存期待","44.0件","残アポ×75%×目標率",GREEN)
metric(d,(850,170,1190,365),"追加期待","80.4件","追加アポ×75%×目標率",ORANGE)
metric(d,(1230,170,1570,365),"合計期待","124.4件","不足124件を上回る",RED)
metric(d,(1610,170,1830,365),"着座率","75%","全員共通前提",NAVY)
card(d,(90,465,1830,890),"変更方針")
text(d,(130,545),"1. 必要着座は全員 75% で固定",31,INK,True)
text(d,(130,615),"2. 目標成約率は、芝隼人・田仲50%、関口・早川45%、その他40%",31,INK,True)
text(d,(130,685),"3. ベンチは最低稼働を維持し、通常メンバーの追加アポ負担を下げる",31,INK,True)
text(d,(130,780),"結論：この前提なら追加246アポで不足124成約をほぼ回収可能。264件で走る場合は上振れバッファ。",32,RED,True)
save(im,14,"target_rate_summary")

# 15 table
im,d=canvas("メンバー別 必要数値：目標成約率版","追加アポ・着座75%・目標成約率で再計算")
card(d,(55,145,1865,940))
headers=["メンバー","区分","目標率","追加アポ","必要着座","追加期待","既存期待","合計期待","日平均"]
xs=[85,265,410,555,735,930,1125,1320,1545]
for x,h in zip(xs,headers): text(d,(x,205),h,21,MUTED,True)
maxadd=max(m["add"] for m in members)
for i,m in enumerate(members):
    y=260+i*58
    color=ORANGE if m["type"]=="ベンチ" else BLUE
    vals=[m["name"],m["type"],f'{m["rate"]*100:.0f}%',f'{m["add"]}件',f'{m["seat"]:.1f}件',f'{m["add_exp"]:.1f}件',f'{m["exist_exp"]:.1f}件',f'{m["total_exp"]:.1f}件',f'{m["add"]/18:.1f}/日']
    for x,v in zip(xs,vals):
        fill=color if x==265 else INK
        text(d,(x,y),v,20,fill,True if x in [85,265,410] else False)
    bar(d,1690,y-2,100,22,m["add"],maxadd,color)
text(d,(90,910),"計算式：期待成約 = アポ数 × 75%着座 × メンバー別目標成約率。",26,RED,True)
save(im,15,"target_rate_member_numbers")

# 16 schedule
im,d=canvas("目標成約率版：日別追加アポスケジュール","合計246件")
x0,y0=48,150
name_w,type_w,date_w,total_w=178,82,78,78
row_h=54
table_w=name_w+type_w+date_w*len(dates)+total_w
table_h=row_h*(len(members)+2)
d.rounded_rectangle([x0-10,y0-12,x0+table_w+10,y0+table_h+10],radius=18,fill=WHITE,outline="#C8D9F3",width=2)
d.rectangle([x0,y0,x0+table_w,y0+row_h],fill="#C8D9F3")
text(d,(x0+12,y0+row_h/2),"メンバー",18,INK,True,"lm")
text(d,(x0+name_w+12,y0+row_h/2),"区分",18,INK,True,"lm")
for i,dt in enumerate(dates):
    x=x0+name_w+type_w+i*date_w
    text(d,(x+date_w/2,y0+row_h/2),dt,16,INK,True,"mm")
text(d,(x0+name_w+type_w+date_w*18+total_w/2,y0+row_h/2),"合計",18,INK,True,"mm")
for r,m in enumerate(members):
    y=y0+row_h*(r+1)
    fill="#FFF8EA" if m["type"]=="ベンチ" else WHITE
    d.rectangle([x0,y,x0+table_w,y+row_h],fill=fill)
    color=ORANGE if m["type"]=="ベンチ" else BLUE
    text(d,(x0+12,y+row_h/2),m["name"],18,INK,True,"lm")
    text(d,(x0+name_w+12,y+row_h/2),m["type"],16,color,True,"lm")
    vals=schedule[m["name"]]
    for i,v in enumerate(vals):
        x=x0+name_w+type_w+i*date_w
        if v:
            d.rounded_rectangle([x+10,y+9,x+date_w-10,y+row_h-9],radius=10,fill=color)
            text(d,(x+date_w/2,y+row_h/2),str(v),18,WHITE,True,"mm")
        else:
            text(d,(x+date_w/2,y+row_h/2),"-",17,"#9AA8B6",False,"mm")
    text(d,(x0+name_w+type_w+date_w*18+total_w/2,y+row_h/2),str(sum(vals)),19,INK,True,"mm")
y=y0+row_h*(len(members)+1)
d.rectangle([x0,y,x0+table_w,y+row_h],fill=YELLOW)
text(d,(x0+12,y+row_h/2),"日計",18,INK,True,"lm")
for i,v in enumerate(daily):
    x=x0+name_w+type_w+i*date_w
    text(d,(x+date_w/2,y+row_h/2),str(v),18,RED if v>=15 else INK,True,"mm")
text(d,(x0+name_w+type_w+date_w*18+total_w/2,y+row_h/2),str(sum(daily)),19,RED,True,"mm")
for c in range(len(dates)+3):
    if c==0: x=x0
    elif c==1: x=x0+name_w
    elif c==2: x=x0+name_w+type_w
    else: x=x0+name_w+type_w+(c-2)*date_w
    d.line([x,y0,x,y0+table_h],fill=LINE,width=1)
for r in range(len(members)+3):
    yy=y0+r*row_h
    d.line([x0,yy,x0+table_w,yy],fill=LINE,width=1)
text(d,(60,990),"ベンチ最低稼働は維持。目標成約率を上げた分、追加アポは246件まで圧縮。",26,RED,True)
save(im,16,"target_rate_daily_schedule")

zip_path=OUT/"orochi_may_revival_presentation_images.zip"
with zipfile.ZipFile(zip_path,"w",zipfile.ZIP_DEFLATED) as z:
    for p in sorted(OUT.glob("*.png")):
        z.write(p,p.name)
print(XLSX)
print(zip_path)
