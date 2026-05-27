from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

path = Path(r"C:\Users\celic\dormswap\outputs\image_excel\orochi_may_revival_presentation.xlsx")
wb = openpyxl.load_workbook(path)
for name in ["Presentation_Targets_75", "Presentation_SeatAction"]:
    if name in wb.sheetnames:
        del wb[name]

members = [
    ("芝隼人","通常",32,1.8,11.2,15.8),
    ("田仲由敬","通常",36,2.0,12.6,22.7),
    ("持木玲那","ベンチ",20,1.1,7.0,7.0),
    ("木原桃香","通常",0,0.0,0.0,0.0),
    ("加藤陸","ベンチ",24,1.3,8.4,13.3),
    ("河上まちこ","ベンチ",20,1.1,7.0,8.0),
    ("五十嵐凌大","ベンチ",24,1.3,8.4,9.1),
    ("早川大貴","通常",22,1.2,7.7,10.1),
    ("笠松佑衣","通常",31,1.7,10.9,17.2),
    ("和佐田舞緒","通常",19,1.1,6.7,10.2),
    ("関口愛里","通常",36,2.0,12.6,24.9),
]

def title(ws, text, cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(1, 1, text)
    c.fill = PatternFill("solid", fgColor="0B3F66")
    c.font = Font(name="Meiryo UI", size=16, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")

def header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row, c)
        cell.fill = PatternFill("solid", fgColor="C8D9F3")
        cell.font = Font(name="Meiryo UI", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style(ws):
    thin = Side(style="thin", color="D9E2EC")
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Meiryo UI", size=10, bold=cell.font.bold, color=cell.font.color)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(top=thin,bottom=thin,left=thin,right=thin)
    for c in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(c)].width = 18

ws = wb.create_sheet("Presentation_Targets_75", 4)
title(ws, "発表用：目標達成に必要なメンバー別数値（着座率75%版）", 10)
ws.append(["メンバー名","区分","追加アポ","日平均","追加期待成約","合計期待","必要着座目安(75%)","着座後必要成約率","優先度","コメント"])
header(ws,2,10)
for m in members:
    name, typ, add, avg, add_exp, total_exp = m
    seat = round(add * 0.75, 1)
    close_req = round(add_exp / seat, 3) if seat else 0
    priority = "最優先" if add >= 31 or typ == "ベンチ" else ("高" if add >= 20 else "通常")
    ws.append([name, typ, add, avg, add_exp, total_exp, seat, close_req, priority, "必要着座は一律75%で算出"])
for r in range(3, 14):
    ws.cell(r, 8).number_format = "0.0%"
style(ws)

ws = wb.create_sheet("Presentation_SeatAction", 5)
title(ws, "発表用：着座率75%にするためのアクションプラン", 7)
ws.append(["タイミング","目的","具体アクション","ツール","担当","完了基準","備考"])
header(ws,2,7)
rows = [
    ["前日", "翌日アポの着座率を上げる", "営業担当に伝えたいことを確認し、顧客ごとの連絡意図を整理", "顧客管理アプリ", "各メンバー", "翌日顧客ごとに連絡内容が決まっている", "名前をタップして直接連絡"],
    ["前日", "返信率を上げる", "ChatGPTで文章を精査し、返信が来やすい文面にする", "ChatGPT", "各メンバー", "送信前に文面チェック済み", "長文ではなく返信しやすくする"],
    ["当日朝", "連絡漏れ防止", "顧客管理アプリから翌日の顧客名をタップして連絡履歴を確認", "顧客管理アプリ", "各メンバー", "対象者全員に朝連絡済み", "手入力検索を避ける"],
    ["当日朝", "着座率75%へ寄せる", "未返信者へ再連絡し、着座見込みを確定", "LINE/連絡ツール", "各メンバー", "翌日アポの75%以上が着座見込み", "日次で田仲さん確認"],
]
for row in rows:
    ws.append(row)
for col in ["B","C","G"]:
    ws.column_dimensions[col].width = 38
for r in range(3, 7):
    ws.row_dimensions[r].height = 50
style(ws)

wb.save(path)
print(path)
