"""
元公務員 Instagram アカウント収集ツール
======================================
APIキー不要・ログイン不要で動作します。
結果はExcelファイルに出力されます。

使い方:
  pip install duckduckgo-search instaloader openpyxl
  python collect.py
"""

import time
import re
import json
from pathlib import Path
from datetime import datetime
import sys
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import instaloader
try:
    from ddgs import DDGS
except ImportError:
    from duckduckgo_search import DDGS

# Windows環境での文字化け対策
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# ==================== 設定 ====================
MIN_FOLLOWERS = 10_000          # フォロワー最小値
MAX_ACCOUNTS  = 100             # 最大収集件数
DELAY_SECONDS = 4               # プロフィール取得間隔（秒）
OUTPUT_FILE   = "元公務員_インスタ.xlsx"
# ==============================================


# ── 検索クエリ ──────────────────────────────
SEARCH_QUERIES = [
    'site:instagram.com "元公務員" 暮らし',
    'site:instagram.com "元公務員" 料理',
    'site:instagram.com "元公務員" 美容',
    'site:instagram.com "元公務員" グルメ',
    'site:instagram.com "元公務員" ファッション',
    'site:instagram.com "元公務員" 子育て',
    'site:instagram.com "元公務員" 旅行',
    'site:instagram.com "元公務員" ヨガ',
    'site:instagram.com "元公務員" ライフスタイル',
    'site:instagram.com "元公務員" ダイエット',
    'site:instagram.com "現役公務員" 暮らし',
    'site:instagram.com "現役公務員" 料理',
    'site:instagram.com "現役公務員" 美容',
    '元公務員 インスタ フォロワー 暮らし ライフスタイル',
    '元公務員 インスタ フォロワー 料理 グルメ',
    '元公務員 インスタ フォロワー 美容 コスメ',
    '元公務員 インスタグラマー 子育て ママ',
    '元公務員 インスタ 旅行 観光',
    '公務員 退職 インスタ 暮らし',
    '公務員 辞めた インスタ 料理',
    '元公務員 主婦 インスタ フォロワー',
]

# ── 公務員キーワード（bioに1つでも含まれればOK）──
CIVIL_SERVANT_KEYWORDS = [
    '元公務員', '前公務員', '公務員を辞め', '公務員卒業', '公務員退職',
    '現役公務員', '現職公務員', '公務員しながら', '公務員×', '元・公務員',
    '公務員→', '公務員出身', '元 公務員', '公務員から',
]

# ── ビジネス系除外キーワード（bioに含まれたら除外）──
EXCLUDE_KEYWORDS = [
    'SNS運用代行', 'SNSマーケター', 'SNSコンサル', 'インスタ講座',
    'SNS講座', '副業スクール', 'コーチング講座', '起業コンサル',
    'マーケティング講師', 'ビジネス塾', '集客コンサル', 'マネタイズ講座',
    'インスタ運用サポート', 'フォロワー獲得', 'インフルエンサー育成',
    '収益化サポート', '副業教室', '月収', '不労所得',
]

# ── ジャンル自動判定（bioのキーワードから）──
GENRE_MAP = [
    ('料理/グルメ',          ['料理', 'レシピ', 'グルメ', '食べ歩き', 'ご飯', '献立', 'おうちごはん', 'cooking', 'food']),
    ('美容/コスメ',          ['美容', 'コスメ', 'スキンケア', 'メイク', 'makeup', 'beauty', 'コスメ好き']),
    ('ファッション',         ['ファッション', 'コーデ', 'fashion', 'outfit', 'style', 'プチプラ']),
    ('子育て/ママ',          ['子育て', 'ママ', '育児', '主婦', '0歳', '1歳', '2歳', '3歳', '小学生ママ']),
    ('旅行',                 ['旅行', 'トラベル', '旅', 'travel', '観光', '一人旅']),
    ('フィットネス/健康',    ['ヨガ', 'フィットネス', 'ダイエット', 'トレーニング', 'yoga', 'fitness', '健康']),
    ('暮らし/ライフスタイル',['暮らし', 'ライフスタイル', 'インテリア', 'シンプルライフ', '丁寧な暮らし', 'life']),
    ('キャリア/働き方',      ['キャリア', '働き方', 'ワーママ', '転職', 'フリーランス', '副業']),
]


def guess_genre(bio: str, fullname: str) -> str:
    text = bio + " " + fullname
    for genre, keywords in GENRE_MAP:
        if any(kw in text for kw in keywords):
            return genre
    return 'その他'


def extract_username(url: str) -> str | None:
    if not url:
        return None
    skip = ['p/', 'reel/', 'reels/', 'explore/', 'accounts/', 'stories/', 'hashtag/']
    if any(s in url for s in skip):
        return None
    m = re.search(r'instagram\.com/([A-Za-z0-9_.]+)/?', url)
    if not m:
        return None
    username = m.group(1)
    if username.lower() in {'p', 'reel', 'reels', 'explore', 'accounts', 'stories',
                             'tv', 'about', 'legal', 'press', 'api', 'directory',
                             'challenge', 'hashtag', 'login', 'signup'}:
        return None
    return username


def search_accounts() -> list[str]:
    """DuckDuckGoで検索してユーザー名リストを返す"""
    found = set()
    print("\n【Step 1】 アカウントを検索中...\n")

    with DDGS() as ddgs:
        for query in SEARCH_QUERIES:
            print(f"  検索: {query}")
            try:
                results = ddgs.text(query, max_results=10, region='jp-jp')
                for r in results:
                    for field in (r.get('href', ''), r.get('body', ''), r.get('title', '')):
                        username = extract_username(field)
                        if username and username not in found:
                            found.add(username)
                time.sleep(1.5)
            except Exception as e:
                print(f"    [警告] {e}")
                time.sleep(3)

    print(f"\n  → {len(found)}件のアカウントを発見\n")
    return list(found)


def fetch_and_filter(usernames: list[str]) -> list[dict]:
    """プロフィールを取得してフィルタリングした結果を返す"""
    print("【Step 2】 各アカウントを確認中...\n")

    loader = instaloader.Instaloader(
        download_pictures=False,
        download_videos=False,
        save_metadata=False,
        quiet=True,
    )

    results = []
    cache_path = Path("cache.json")
    cache = json.loads(cache_path.read_text(encoding='utf-8')) if cache_path.exists() else {"done": []}
    done_set = set(cache["done"])

    for i, username in enumerate(usernames):
        if len(results) >= MAX_ACCOUNTS:
            print(f"  最大件数（{MAX_ACCOUNTS}件）に達しました。")
            break

        if username in done_set:
            continue

        print(f"  [{i+1}/{len(usernames)}] @{username}")

        try:
            profile = instaloader.Profile.from_username(loader.context, username)
            time.sleep(DELAY_SECONDS)
        except instaloader.exceptions.TooManyRequestsException:
            print("    [待機] レート制限。60秒待機...")
            time.sleep(60)
            try:
                profile = instaloader.Profile.from_username(loader.context, username)
            except Exception:
                continue
        except Exception as e:
            print(f"    [スキップ] {e}")
            done_set.add(username)
            continue

        bio = profile.biography or ''
        followers = profile.followers

        # フォロワー数チェック
        if followers < MIN_FOLLOWERS:
            print(f"    × フォロワー不足 ({followers:,}人)")
            done_set.add(username)
            continue

        # 公務員キーワードチェック
        has_civil = any(kw in bio or kw in profile.full_name for kw in CIVIL_SERVANT_KEYWORDS)
        if not has_civil:
            print(f"    × 公務員キーワードなし")
            done_set.add(username)
            continue

        # ビジネス系除外
        excluded_kw = next((kw for kw in EXCLUDE_KEYWORDS if kw in bio), None)
        if excluded_kw:
            print(f"    × ビジネス系除外: {excluded_kw}")
            done_set.add(username)
            continue

        # 通過！
        genre = guess_genre(bio, profile.full_name)
        followers_man = followers // 10000  # 万単位
        display_name = profile.full_name or username
        feature = f"{next((kw for kw in CIVIL_SERVANT_KEYWORDS if kw in bio or kw in profile.full_name), '元公務員')}・{bio[:40].replace(chr(10), '・').strip('・')}・{followers_man}万F"

        results.append({
            'username':     username,
            'display_name': display_name,
            'followers':    followers,
            'bio':          bio,
            'genre':        genre,
            'feature':      feature,
            'url':          f"https://www.instagram.com/{username}/",
        })
        print(f"    ✓ {display_name} | {genre} | {followers:,}F")

        done_set.add(username)

        # キャッシュ保存（5件ごと）
        if (i + 1) % 5 == 0:
            cache["done"] = list(done_set)
            cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding='utf-8')

    cache["done"] = list(done_set)
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding='utf-8')

    print(f"\n  → 条件に合致: {len(results)}件\n")
    return results


def save_to_excel(accounts: list[dict]):
    """Excelファイルに出力する（スプレッドシートに貼るだけの形式）"""
    if not accounts:
        print("保存するデータがありません。")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "元公務員"

    # ヘッダー
    headers = ['No.', 'アカウント名', 'ジャンル(メイン)', '特徴(サブ)', 'URL',
               'フォロワー数', 'bio（参考）']
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # データ行
    for i, acc in enumerate(accounts, 1):
        row = [
            i,
            acc['display_name'],
            acc['genre'],
            acc['feature'],
            acc['url'],
            acc['followers'],
            acc['bio'],
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            if col == 5:  # URLはハイパーリンク
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")
            cell.alignment = Alignment(wrap_text=True)

        # 行の背景色（交互）
        if i % 2 == 0:
            fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
            for col in range(1, 8):
                ws.cell(row=i + 1, column=col).fill = fill

    # 列幅調整
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 60

    ws.row_dimensions[1].height = 20

    wb.save(OUTPUT_FILE)
    print(f"【完了】 {OUTPUT_FILE} に {len(accounts)}件を保存しました")
    print(f"  → A〜E列（No.〜URL）をコピーしてスプレッドシートに貼り付けてください")


def main():
    print("=" * 55)
    print(" 元公務員 Instagram アカウント収集ツール")
    print("=" * 55)

    # Step 1: 検索
    usernames = search_accounts()

    if not usernames:
        print("アカウントが見つかりませんでした。")
        return

    # Step 2: プロフィール取得＆フィルター
    accounts = fetch_and_filter(usernames)

    # Step 3: Excel出力
    print("【Step 3】 Excelファイルを出力中...")
    save_to_excel(accounts)


if __name__ == "__main__":
    main()
