from pathlib import Path
import shutil

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


BASE = Path(r"C:\Users\celic\dormswap")
OUT = BASE / "generated_household"
OUT.mkdir(exist_ok=True)

FONT_PATH = Path(r"C:\Windows\Fonts\YuGothR.ttc")
FONT_BOLD_PATH = Path(r"C:\Windows\Fonts\YuGothB.ttc")
if FONT_PATH.exists():
    pdfmetrics.registerFont(TTFont("YuGothic", str(FONT_PATH)))
    pdfmetrics.registerFont(TTFont("YuGothic-Bold", str(FONT_BOLD_PATH if FONT_BOLD_PATH.exists() else FONT_PATH)))
    BASE_FONT = "YuGothic"
    BOLD_FONT = "YuGothic-Bold"
else:
    BASE_FONT = "Helvetica"
    BOLD_FONT = "Helvetica-Bold"


MONTHS = [
    {
        "label": "2026年2月",
        "sheet": "2026-02",
        "salary": 252_500,
        "income_total": 450_291,
        "expense_total": 464_264,
        "comm": 0,
        "health": 0,
        "fee": 0,
        "notes": [
            "給与は日研トータルソーシング 2026年1月分給与明細（2026/02/20入金）を根拠。",
            "ゆうちょ銀行は本人申告により取引なし。",
            "デビット明細の店舗名までは未取得のため、細目はSBI明細上で判読できる範囲。",
        ],
    },
    {
        "label": "2026年3月",
        "sheet": "2026-03",
        "salary": 265_493,
        "income_total": 496_819,
        "expense_total": 492_090,
        "comm": 0,
        "health": 0,
        "fee": 0,
        "notes": [
            "給与は日研トータルソーシング 2026年2月分給与明細（2026/03/19入金）を根拠。",
            "ゆうちょ銀行は本人申告により取引なし。",
            "3月分SBI PDFは3/1-3/31を取得済み。あまた法律事務所の要求期間3/15以降を含む。",
        ],
    },
    {
        "label": "2026年4月",
        "sheet": "2026-04",
        "salary": 247_668,
        "income_total": 383_420,
        "expense_total": 383_306,
        "comm": 6_541,
        "health": 44_968,
        "fee": 7_678,
        "notes": [
            "給与は住信SBI明細上の「給与＊ニツケントータルソーシング（カ）」247,668円を根拠。",
            "Levelaの4月報酬請求615,384円は、住信SBI上の入金が2026/05/16のため4月入金給与とは分けて記載。",
            "ゆうちょ銀行は本人申告により取引なし。",
        ],
    },
]

for month in MONTHS:
    month["other_income"] = month["income_total"] - month["salary"]
    month["living_other"] = month["expense_total"] - month["comm"] - month["health"] - month["fee"]


def yen(value: int) -> str:
    return f"{value:,}円"


def build_workbook() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    headers = [
        "月",
        "給与・報酬等",
        "その他入金・口座内振替等",
        "SBI入金合計",
        "通信費",
        "健康保険等",
        "会費等",
        "その他生活費・デビット・現金等",
        "SBI出金合計",
        "収支差額",
        "備考",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style="thin", color="AAAAAA")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for month in MONTHS:
        ws.append(
            [
                month["label"],
                month["salary"],
                month["other_income"],
                month["income_total"],
                month["comm"],
                month["health"],
                month["fee"],
                month["living_other"],
                month["expense_total"],
                month["income_total"] - month["expense_total"],
                "SBI明細から確認できる範囲。ゆうちょ取引なし。",
            ]
        )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=10):
        for cell in row:
            cell.number_format = "#,##0"
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["K"].width = 45

    for month in MONTHS:
        wsm = wb.create_sheet(month["sheet"])
        rows = [
            ["項目", "金額", "根拠・メモ"],
            ["給与・報酬等", month["salary"], "給与明細または住信SBI明細"],
            ["その他入金・口座内振替等", month["other_income"], "SBI入金合計から給与・報酬等を差引"],
            ["入金合計", month["income_total"], "住信SBI月別合計"],
            ["通信費", month["comm"], "4月はドコモ ケイタイを確認。その他月はSBIのみでは未分類"],
            ["健康保険等", month["health"], "4月は日研グループ健康保険組合への振込を確認"],
            ["会費等", month["fee"], "4月はエニタイム会費を確認"],
            ["その他生活費・デビット・現金等", month["living_other"], "デビット、ATM、PayPay、口座内振替等を含む未分類分"],
            ["出金合計", month["expense_total"], "住信SBI月別合計"],
            ["収支差額", month["income_total"] - month["expense_total"], "入金合計 - 出金合計"],
        ]
        for row in rows:
            wsm.append(row)
        for cell in wsm[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for row in wsm.iter_rows():
            for cell in row:
                cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for idx, width in enumerate([28, 14, 55], start=1):
            wsm.column_dimensions[get_column_letter(idx)].width = width
        for row in wsm.iter_rows(min_row=2, min_col=2, max_col=2):
            row[0].number_format = "#,##0"
        start = wsm.max_row + 2
        wsm.cell(start, 1, "注記").font = Font(bold=True)
        for row_idx, note in enumerate(month["notes"], start + 1):
            wsm.cell(row_idx, 1, note)
            wsm.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)

    path = OUT / "家計表_SBI明細ベース_2026年02月-04月.xlsx"
    wb.save(path)
    return path


def table_for_month(month: dict) -> Table:
    data = [
        ["区分", "項目", "金額", "根拠・メモ"],
        ["収入", "給与・報酬等", yen(month["salary"]), "給与明細またはSBI明細"],
        ["収入", "その他入金・口座内振替等", yen(month["other_income"]), "SBI入金合計から給与・報酬等を差引"],
        ["収入", "入金合計", yen(month["income_total"]), "住信SBI月別合計"],
        ["支出", "通信費", yen(month["comm"]), "4月はドコモ ケイタイを確認。その他月は未分類"],
        ["支出", "健康保険等", yen(month["health"]), "4月は日研グループ健康保険組合への振込を確認"],
        ["支出", "会費等", yen(month["fee"]), "4月はエニタイム会費を確認"],
        ["支出", "その他生活費・デビット・現金等", yen(month["living_other"]), "デビット、ATM、PayPay、口座内振替等を含む未分類分"],
        ["支出", "出金合計", yen(month["expense_total"]), "住信SBI月別合計"],
        ["差額", "収支差額", yen(month["income_total"] - month["expense_total"]), "入金合計 - 出金合計"],
    ]
    table = Table(data, colWidths=[18 * mm, 50 * mm, 28 * mm, 78 * mm])
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), BASE_FONT),
                ("FONTNAME", (0, 0), (-1, 0), BOLD_FONT),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (2, 1), (2, -1), "RIGHT"),
                ("BACKGROUND", (0, 3), (-1, 3), colors.HexColor("#D9EAF7")),
                ("BACKGROUND", (0, 8), (-1, 8), colors.HexColor("#FCE4D6")),
            ]
        )
    )
    return table


def build_pdf(path: Path, months: list[dict]) -> None:
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="JPTitle", fontName=BOLD_FONT, fontSize=18, leading=24, alignment=1, spaceAfter=8))
    styles.add(ParagraphStyle(name="JP", fontName=BASE_FONT, fontSize=9, leading=13))
    styles.add(ParagraphStyle(name="JPBold", fontName=BOLD_FONT, fontSize=10, leading=14))
    doc = SimpleDocTemplate(
        str(path), pagesize=A4, rightMargin=14 * mm, leftMargin=14 * mm, topMargin=12 * mm, bottomMargin=12 * mm
    )
    story = []
    for idx, month in enumerate(months):
        story.append(Paragraph(f"家計表（SBI明細ベース） {month['label']}", styles["JPTitle"]))
        story.append(
            Paragraph(
                "作成根拠: 住信SBIネット銀行入出金明細、給与明細。ゆうちょ銀行は本人申告により取引なし。"
                "デビット利用先の詳細店名が未取得のため、判読できない支出は「その他生活費・デビット・現金等」に集約しています。",
                styles["JP"],
            )
        )
        story.append(Spacer(1, 5 * mm))
        story.append(table_for_month(month))
        story.append(Spacer(1, 5 * mm))
        story.append(Paragraph("注記", styles["JPBold"]))
        for note in month["notes"]:
            story.append(Paragraph("・" + note, styles["JP"]))
        if idx != len(months) - 1:
            story.append(PageBreak())
    doc.build(story)


def main() -> None:
    workbook = build_workbook()
    combined_pdf = OUT / "家計表_SBI明細ベース_2026年02月-04月.pdf"
    build_pdf(combined_pdf, MONTHS)
    for month in MONTHS:
        build_pdf(OUT / f"家計表_SBI明細ベース_{month['label']}.pdf", [month])

    for dest in [
        Path(r"C:\Users\celic\Desktop\弁護士整理\08_家計表"),
        Path(r"C:\Users\celic\OneDrive\デスクトップ\弁護士整理\08_家計表"),
    ]:
        dest.mkdir(parents=True, exist_ok=True)
        for file in OUT.iterdir():
            if file.name.startswith("家計表_SBI明細ベース"):
                shutil.copy2(file, dest / file.name)

    print("created")
    for file in sorted(OUT.iterdir()):
        print(file.name, file.stat().st_size)


if __name__ == "__main__":
    main()
